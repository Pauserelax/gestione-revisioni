"""Interfaccia a riga di comando: import, scadenzario, export liste chiamata."""

from __future__ import annotations

import argparse
import sys
from datetime import date
from pathlib import Path

from . import db as database
from .esporta import esporta_lista_chiamate
from .parser_ccf import leggi_report_ccf
from .parser_tcar import leggi_export_lead
from .scadenze import calcola_scadenze, da_contattare

BASE = Path(__file__).resolve().parent.parent.parent
DB_PREDEFINITO = BASE / "dati" / "revisioni.db"


def _stampa_tabella(scadenze, limite=None) -> None:
    fmt = "{:<12} {:<11} {:>6}  {:<30} {:<16} {:<5} {:<8} {:<10}"
    print(fmt.format("STATO", "SCADENZA", "GG", "CLIENTE", "TELEFONO", "MARCA", "MODELLO", "FONTE"))
    print("-" * 110)
    for s in scadenze[:limite]:
        print(fmt.format(
            s.stato,
            s.scadenza.strftime("%d/%m/%Y") if s.scadenza else "-",
            s.giorni_rimanenti if s.giorni_rimanenti is not None else "-",
            s.cliente[:30],
            s.telefono[:16],
            s.marca,
            s.modello,
            "presunta" if s.fonte_data == "mese_report" else s.fonte_data,
        ))
    if limite and len(scadenze) > limite:
        print(f"... e altri {len(scadenze) - limite} veicoli")


def cmd_importa(args) -> None:
    percorso = Path(args.file)
    if not percorso.exists():
        sys.exit(f"File non trovato: {percorso}")
    righe, punti_vendita = leggi_report_ccf(percorso)
    if not righe:
        sys.exit("Nessuna riga veicolo riconosciuta: il file non sembra un riepilogo C.C.F. di WinDrakkar.")
    conn = database.apri_db(Path(args.db))
    esito = database.importa_righe(conn, righe, percorso.name, punti_vendita)
    print(f"Import di {percorso.name}: {esito['righe']} righe lette, "
          f"{esito['nuovi']} veicoli nuovi, {esito['aggiornati']} aggiornati.")


def cmd_importa_storico(args) -> None:
    from .parser_storico import leggi_file_storico, trova_file_storici
    percorso = Path(args.percorso)
    if not percorso.exists():
        sys.exit(f"Percorso non trovato: {percorso}")
    files = trova_file_storici(percorso) if percorso.is_dir() else [percorso]
    if not files:
        sys.exit("Nessun file .xlsx/.xls trovato.")
    conn = database.apri_db(Path(args.db))
    totali = {"righe": 0, "nuovi": 0, "aggiornati": 0, "revisioni": 0, "feedback": 0, "dismessi": 0}
    errori = []
    for i, f in enumerate(files, 1):
        try:
            righe = leggi_file_storico(f)
        except Exception as e:
            errori.append(f"{f.name}: {e}")
            print(f"[{i}/{len(files)}] {f.name}: ERRORE lettura ({str(e)[:60]})")
            continue
        if not righe:
            print(f"[{i}/{len(files)}] {f.name}: nessuna riga riconosciuta, saltato")
            continue
        esito = database.importa_storico(conn, righe, f.name)
        totali = {k: totali[k] + esito.get(k, 0) for k in totali}
        print(f"[{i}/{len(files)}] {f.name}: {esito['righe']} righe, "
              f"{esito['nuovi']} veicoli nuovi, {esito['revisioni']} revisioni, {esito['feedback']} esiti")
    print("\n=== TOTALE IMPORT STORICO ===")
    print(f"File elaborati: {len(files) - len(errori)}/{len(files)}")
    print(f"Righe lette: {totali['righe']} — Veicoli nuovi: {totali['nuovi']} — "
          f"aggiornamenti: {totali['aggiornati']}")
    print(f"Revisioni registrate: {totali['revisioni']} — Esiti storici: {totali['feedback']} — "
          f"Veicoli dismessi (VENDUTA/DEMOLITA): {totali['dismessi']}")
    if errori:
        print("\nFile con errori:")
        for e in errori:
            print("  -", e)


def cmd_importa_immatricolazioni(args) -> None:
    import csv

    from .parser_immatricolazioni import leggi_file_immatricolazioni, trova_file
    percorso = Path(args.percorso)
    files = trova_file(percorso) if percorso.is_dir() else [percorso]
    if not files:
        sys.exit("Nessun file .xlsx trovato.")
    conn = database.apri_db(Path(args.db))
    totale = {"righe": 0, "match_telaio": 0, "match_targa": 0, "match_cliente": 0,
              "nuovi": 0, "date_aggiornate": 0, "targhe_aggiunte": 0}
    divergenze = []
    for i, f in enumerate(files, 1):
        righe = leggi_file_immatricolazioni(f)
        if not righe:
            print(f"[{i}/{len(files)}] {f.name}: saltato (mese non riconosciuto o vuoto)")
            continue
        esito = database.importa_immatricolazioni(conn, righe, f.name)
        divergenze.extend(esito["divergenze"])
        for k in totale:
            totale[k] += esito[k]
        print(f"[{i}/{len(files)}] {f.name}: {esito['righe']} righe, "
              f"{esito['date_aggiornate']} date aggiornate, {esito['nuovi']} veicoli nuovi")
    print("\n=== TOTALE IMMATRICOLAZIONI ===")
    print(f"Righe: {totale['righe']} — match per telaio {totale['match_telaio']}, "
          f"per targa {totale['match_targa']}, per cliente {totale['match_cliente']}, "
          f"veicoli nuovi {totale['nuovi']}")
    print(f"Date immatricolazione aggiornate: {totale['date_aggiornate']} — "
          f"targhe aggiunte a veicoli che non l'avevano: {totale['targhe_aggiunte']}")
    if divergenze:
        out = BASE / "liste" / f"divergenze_date_{date.today().isoformat()}.csv"
        with open(out, "w", newline="", encoding="utf-8-sig") as fh:
            w = csv.DictWriter(fh, fieldnames=list(divergenze[0].keys()), delimiter=";")
            w.writeheader()
            w.writerows(divergenze)
        print(f"Divergenze tra data garanzia/Tcar e immatricolazioni: {len(divergenze)} → {out}")
    stimate = database.azzera_date_stimate(conn)
    if stimate:
        print(f"Date 'stimate' residue azzerate: {stimate} (meglio nessuna data che una sbagliata).")


def cmd_importa_dekra(args) -> None:
    from .parser_dekra import leggi_pdf_dekra, trova_file_dekra
    percorso = Path(args.percorso)
    if not percorso.exists():
        sys.exit(f"Percorso non trovato: {percorso}")
    files = trova_file_dekra(percorso) if percorso.is_dir() else [percorso]
    if not files:
        sys.exit("Nessun file .pdf trovato.")
    conn = database.apri_db(Path(args.db))
    totali = {"righe": 0, "nuove": 0, "doppie": 0, "revisioni": 0}
    esito = None
    for i, f in enumerate(files, 1):
        try:
            righe = leggi_pdf_dekra(f)
        except ImportError:
            sys.exit("Manca una libreria PDF (pypdf o pdfplumber): installala con  python -m pip install pypdf")
        if not righe:
            print(f"[{i}/{len(files)}] {f.name}: nessuna riga riconosciuta, saltato")
            continue
        esito = database.importa_dekra(conn, righe, f.name)
        for k in totali:
            totali[k] += esito[k]
        print(f"[{i}/{len(files)}] {f.name}: {esito['righe']} passaggi letti, "
              f"{esito['nuove']} nuovi, {esito['doppie']} già presenti")
    if esito is None:
        sys.exit("Nessun PDF Dekra riconosciuto.")
    print("\n=== TOTALE DEKRA ===")
    print(f"Passaggi in linea: {totali['righe']} letti, {totali['nuove']} nuovi registrati")
    print(f"Revisioni REGOLARI aggiunte allo storico veicoli: {totali['revisioni']}")
    print(f"Passaggi agganciati al parco: righe con veicolo nostro {esito['match']}, "
          f"riagganci di vecchi orfani {esito['riagganciate']}")
    print(f"Targhe/telai completati sui veicoli: {esito['targhe_aggiunte']}/{esito['telai_aggiunti']}")
    print(f"Passaggi NON agganciati (veicolo fuori parco): {esito['non_agganciate']}")
    print(f"Clienti abituali della linea (2+ passaggi): {esito['abituali']} targhe")
    if esito["ripristinati"]:
        print(f"Veicoli archiviati tornati in gestione (ora hanno dati validi): {esito['ripristinati']}")


def cmd_importa_tcar(args) -> None:
    percorso = Path(args.file)
    if not percorso.exists():
        sys.exit(f"File non trovato: {percorso}")
    leads = leggi_export_lead(percorso)
    if not leads:
        sys.exit("Nessun lead riconosciuto: il file non sembra un export lead di Tcar.")
    conn = database.apri_db(Path(args.db))
    esito = database.importa_lead_tcar(conn, leads, percorso.name)
    print(f"Import di {percorso.name}: {esito['lead']} lead letti, "
          f"{esito['match']} agganciati a veicoli nostri, "
          f"{esito['arricchiti']} veicoli arricchiti (targa/data immatricolazione).")


def cmd_lead(args) -> None:
    conn = database.apri_db(Path(args.db))
    filtro = "" if args.tutti else "WHERE l.veicolo_id IS NOT NULL"
    righe = conn.execute(f"""
        SELECT l.cognome, l.nome, l.telaio, l.targa, l.stato, l.assegnatario, l.campagna,
               l.veicolo_id, c.nome AS cliente_nostro
        FROM lead_tcar l
        LEFT JOIN veicoli v ON v.id = l.veicolo_id
        LEFT JOIN clienti c ON c.id = v.cliente_id
        {filtro} ORDER BY l.creazione DESC""").fetchall()
    if not righe:
        print("Nessun lead" + ("" if args.tutti else " agganciato a veicoli nostri") + ".")
        return
    for r in righe:
        aggancio = f"NOSTRO CLIENTE: {r['cliente_nostro']}" if r["veicolo_id"] else "non in database"
        print(f"{(r['cognome'] + ' ' + (r['nome'] or '')).strip():<32} {r['targa']:<9} {r['telaio']:<19} "
              f"{r['stato']:<10} {aggancio}")
    print(f"\nTotale: {len(righe)} lead")


def cmd_scadenze(args) -> None:
    conn = database.apri_db(Path(args.db))
    scadenze = calcola_scadenze(conn, giorni_allarme=args.giorni,
                                includi_esclusi=args.includi_esclusi)
    if args.mese:
        lista = sorted((s for s in scadenze
                        if s.scadenza and s.scadenza.strftime("%Y-%m") == args.mese
                        and s.ultimo_esito not in ("appuntamento", "non_interessato")),
                       key=lambda s: (s.scadenza, s.cliente))
        clienti = len({s.cliente for s in lista})
        print(f"Scadenze revisione di {args.mese}: {len(lista)} veicoli di {clienti} clienti\n")
    else:
        lista = scadenze if args.tutte else da_contattare(scadenze)
    if not lista:
        print("Nessun veicolo da contattare nel periodo indicato.")
        return
    _stampa_tabella(lista, limite=args.limite)
    print(f"\nTotale: {len(lista)} veicoli "
          f"({sum(1 for s in lista if s.stato == 'SCADUTA')} scaduti, "
          f"{sum(1 for s in lista if s.stato == 'IN_SCADENZA')} in scadenza questo mese, "
          f"{sum(1 for s in lista if s.stato == 'PROSSIMA')} entro {args.giorni} giorni)")


def cmd_excel(args) -> None:
    conn = database.apri_db(Path(args.db))
    scadenze = calcola_scadenze(conn, giorni_allarme=args.giorni,
                                includi_esclusi=args.includi_esclusi)
    lista = scadenze if args.tutte else da_contattare(scadenze)
    out = Path(args.out) if args.out else BASE / "liste" / f"lista_chiamate_{date.today().isoformat()}.xlsx"
    esporta_lista_chiamate(lista, out)
    print(f"Lista chiamate salvata in: {out} ({len(lista)} veicoli)")


def cmd_esito(args) -> None:
    conn = database.apri_db(Path(args.db))
    row = conn.execute(
        "SELECT v.id, v.telaio, v.targa, v.cliente_id, c.nome FROM veicoli v JOIN clienti c ON c.id = v.cliente_id WHERE v.telaio = ? OR v.targa = ?",
        (args.veicolo.upper(), args.veicolo.upper()),
    ).fetchone()
    if not row:
        sys.exit(f"Veicolo non trovato (telaio o targa): {args.veicolo}")
    scadenze = {s.veicolo_id: s for s in calcola_scadenze(conn, includi_esclusi=True)}
    s = scadenze.get(row["id"])
    scadenza_iso = s.scadenza.isoformat() if s and s.scadenza else ""

    if args.esito == "revisione_fatta":
        data_rev = date.fromisoformat(args.data) if args.data else date.today()
        database.registra_revisione(conn, row["id"], data_rev)
        database.registra_contatto(conn, row["id"], scadenza_iso, args.esito, args.note or "")
        nuova = calcola_scadenze(conn, includi_esclusi=True)
        prossima = next((x.scadenza for x in nuova if x.veicolo_id == row["id"]), None)
        print(f"Revisione del {data_rev.strftime('%d/%m/%Y')} registrata per {row['nome']} ({row['telaio']}).")
        if prossima:
            print(f"Prossima chiamata riprogrammata per la scadenza del {prossima.strftime('%d/%m/%Y')}.")
    elif args.esito == "non_possiede_piu":
        database.dismetti_veicolo(conn, row["id"])
        database.registra_contatto(conn, row["id"], scadenza_iso, args.esito, args.note or "")
        print(f"Veicolo {row['telaio'] or row['targa']} dismesso: {row['nome']} non lo possiede più. "
              f"Escluso da scadenzario e liste (lo storico resta).")
        attivi = conn.execute(
            "SELECT COUNT(*) AS n FROM veicoli WHERE cliente_id = ? AND attivo = 1",
            (row["cliente_id"],),
        ).fetchone()["n"]
        if attivi == 0:
            print(f"⚠ {row['nome']} è rimasto SENZA veicoli in scadenzario: chiedere che auto "
                  f"guida ora e registrarla, restiamo il suo fornitore di revisione. Esempio:\n"
                  f"  python3 -m revisioni nuovo-veicolo \"{row['nome']}\" --targa XX000XX "
                  f"--marca ... --modello ... --imm AAAA-MM-GG\n"
                  f"Il cliente compare nella lista 'da-recuperare' finché non ha un veicolo attivo.")
    else:
        if not scadenza_iso:
            sys.exit("Il veicolo non ha una scadenza calcolabile.")
        database.registra_contatto(conn, row["id"], scadenza_iso, args.esito, args.note or "")
        print(f"Registrato esito '{args.esito}' per {row['nome']} ({row['telaio']}), "
              f"scadenza {s.scadenza.strftime('%d/%m/%Y')}.")


def cmd_nuovo_veicolo(args) -> None:
    conn = database.apri_db(Path(args.db))
    if not args.targa and not args.telaio:
        sys.exit("Serve almeno la targa (--targa) o il telaio (--telaio).")
    if not args.imm and not args.ultima_revisione:
        sys.exit("Serve la data di immatricolazione (--imm) o dell'ultima revisione (--ultima-revisione), "
                 "altrimenti la scadenza non è calcolabile.")

    filtro = f"%{args.cliente}%"
    clienti = conn.execute(
        "SELECT * FROM clienti WHERE nome LIKE ? OR telefono LIKE ?", (filtro, filtro)
    ).fetchall()
    if len(clienti) > 1:
        for c in clienti:
            print(f"  {c['nome']}  tel: {c['telefono'] or '-'}")
        sys.exit(f"Più clienti corrispondono a '{args.cliente}': restringi la ricerca (anche col telefono).")
    if clienti:
        cliente = clienti[0]
        cliente_id = cliente["id"]
        print(f"Cliente esistente: {cliente['nome']} (lo storico prosegue)")
    else:
        if not args.telefono:
            sys.exit(f"Cliente '{args.cliente}' non trovato: per crearlo nuovo indica anche --telefono.")
        cliente_id = conn.execute(
            "INSERT INTO clienti (nome, telefono) VALUES (?, ?)",
            (args.cliente.upper(), args.telefono),
        ).lastrowid
        conn.commit()
        print(f"Cliente nuovo creato: {args.cliente.upper()}")

    imm = date.fromisoformat(args.imm) if args.imm else None
    veicolo_id = database.inserisci_veicolo_manuale(
        conn, cliente_id, (args.targa or "").upper() or None, (args.telaio or "").upper() or None,
        args.marca or "", args.modello or "", imm)
    if args.ultima_revisione:
        database.registra_revisione(conn, veicolo_id, date.fromisoformat(args.ultima_revisione), fonte="dichiarata dal cliente")
    s = next((x for x in calcola_scadenze(conn, includi_esclusi=True) if x.veicolo_id == veicolo_id), None)
    print(f"Veicolo registrato: {args.marca or ''} {args.modello or ''} targa {args.targa or '-'}".strip())
    if s and s.scadenza:
        print(f"In scadenzario: prossima revisione entro il {s.scadenza.strftime('%d/%m/%Y')}.")


def cmd_da_recuperare(args) -> None:
    conn = database.apri_db(Path(args.db))
    righe = database.clienti_senza_veicolo(conn)
    if not righe:
        print("Nessun cliente da recuperare: tutti hanno almeno un veicolo in scadenzario.")
        return
    print("Clienti storici senza veicolo in scadenzario (da richiamare: che auto guidano ora?):\n")
    for r in righe:
        dettagli = " — ".join(filter(None, [
            r["motivo"] and f"motivo: {r['motivo']}",
            r["dismesso_il"] and f"registrato il {r['dismesso_il']}",
            r["fonte"] and f"fonte: {r['fonte']}",
        ]))
        identificativo = " ".join(filter(None, [r["targa"], r["telaio"] and f"({r['telaio']})"]))
        print(f"  {r['nome']:<32} tel: {r['telefono'] or '-':<16} "
              f"veicolo: {r['veicolo'] or '-':<22} {identificativo}"
              + (f"\n{'':36}{dettagli}" if dettagli else ""))
    print(f"\nTotale: {len(righe)} clienti")


def cmd_cliente(args) -> None:
    conn = database.apri_db(Path(args.db))
    filtro = f"%{args.ricerca}%"
    clienti = conn.execute(
        "SELECT * FROM clienti WHERE nome LIKE ? OR telefono LIKE ? ORDER BY nome", (filtro, filtro)
    ).fetchall()
    if not clienti:
        print(f"Nessun cliente trovato per '{args.ricerca}'.")
        return
    scadenze = {s.veicolo_id: s for s in calcola_scadenze(conn, includi_esclusi=True)}
    for c in clienti:
        print(f"\n{c['nome']}  tel: {c['telefono'] or '-'}  email: {c['email'] or '-'}")
        veicoli = conn.execute(
            "SELECT * FROM veicoli WHERE cliente_id = ? ORDER BY data_immatricolazione", (c["id"],)
        ).fetchall()
        for v in veicoli:
            s = scadenze.get(v["id"])
            stato_v = "ARCHIVIATO" if v["archiviato"] else ("attivo" if v["attivo"] else "DISMESSO")
            scad = f"prossima revisione {s.scadenza.strftime('%d/%m/%Y')}" if s and s.scadenza else "scadenza non calcolabile"
            if not v["attivo"]:
                scad = "fuori scadenzario"
            print(f"  [{stato_v}] {v['marca']} {v['modello']} targa {v['targa'] or '-'} telaio {v['telaio'] or '-'}"
                  f" imm. {v['data_immatricolazione'] or '-'} ({v['fonte_data']}) — {scad}")
            for r in conn.execute(
                "SELECT data_revisione, fonte FROM revisioni_effettuate WHERE veicolo_id = ? ORDER BY data_revisione", (v["id"],)
            ):
                print(f"      revisione effettuata il {r['data_revisione']} ({r['fonte']})")
            for k in conn.execute(
                "SELECT data_contatto, esito, note FROM contatti WHERE veicolo_id = ? ORDER BY data_contatto DESC LIMIT 5", (v["id"],)
            ):
                nota = f" — {k['note']}" if k["note"] else ""
                print(f"      contatto {k['data_contatto'][:10]}: {k['esito']}{nota}")


def cmd_storico(args) -> None:
    from .storico_html import genera_storico
    conn = database.apri_db(Path(args.db))
    out = Path(args.out) if args.out else BASE / "liste" / "storico_clienti.html"
    genera_storico(conn, out)
    print(f"Storico clienti navigabile generato: {out}\nAprilo con doppio click (si usa dal browser).")


def cmd_sms(args) -> None:
    from .sms import (TESTO_ARRETRATO_PREDEFINITO, TESTO_PREDEFINITO, _testo_template,
                      esporta_smscafe, lista_sms, segna_inviati, testo_campagna_smscafe)
    conn = database.apri_db(Path(args.db))
    try:
        righe = lista_sms(conn, arretrato=args.arretrato, base=BASE)
    except ValueError as e:
        sys.exit(str(e))
    if not righe:
        print("Nessun SMS da inviare" + (" per l'arretrato." if args.arretrato else ": nessuna scadenza tra 2 mesi senza SMS già inviato."))
        return
    suffisso = "arretrato" if args.arretrato else "scadenze"
    out = Path(args.out) if args.out else BASE / "liste" / f"sms_{suffisso}_{date.today().isoformat()}.xlsx"
    if args.arretrato:
        template = _testo_template(BASE / "dati" / "sms_testo_arretrato.txt", TESTO_ARRETRATO_PREDEFINITO)
    else:
        template = _testo_template(BASE / "dati" / "sms_testo.txt", TESTO_PREDEFINITO)
    esporta_smscafe(righe, out, template)
    print(f"Lista SMS salvata in: {out} ({len(righe)} destinatari, un SMS per numero)")
    print("Formato SMS Café: 1ª riga etichette = SÌ; mappa Nome=Colonna A, Cognome=Colonna B, Cellulare=Colonna C.")
    testo = testo_campagna_smscafe(template, righe)
    print("\nTesto campagna da incollare in SMS Café (lunghezze garantite ≤160: i nomi")
    print("troppo lunghi sono già accorciati nella colonna Nome del file):")
    print(f"  {testo}")
    print("Caricala su smscafè. Esempio primo messaggio:")
    print(f"  → {righe[0]['telefono']}: {righe[0]['messaggio']}")
    if args.segna:
        n = segna_inviati(conn, righe)
        print(f"Registrato esito 'sms_inviato' su {n} veicoli: non verranno riestratti.")
    else:
        print("Quando li hai inviati, rilancia con --segna per registrarli (o usa il bottone in dashboard).")


def cmd_telefoni(args) -> None:
    from collections import Counter

    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    from .telefoni import telefoni_condivisi, valuta_campo
    conn = database.apri_db(Path(args.db))
    clienti = conn.execute("SELECT id, nome, telefono, email FROM clienti ORDER BY nome").fetchall()
    valutati = [(c, *valuta_campo(c["telefono"] or "")) for c in clienti]
    conteggi = Counter(v[1] for v in valutati)
    print("Qualità dei telefoni in anagrafica:")
    etichette = {"cellulare": "Cellulari validi", "fisso": "Fissi validi (no SMS)",
                 "sospetto": "Sospetti/fittizi", "non_valido": "Non validi", "mancante": "Mancanti"}
    for stato in ("cellulare", "fisso", "sospetto", "non_valido", "mancante"):
        print(f"  {etichette[stato]:<24} {conteggi.get(stato, 0):>6}")
    condivisi = telefoni_condivisi(conn)
    print(f"  {'Numeri condivisi da 3+ clienti':<24} {len(condivisi):>6}")
    for r in condivisi[:5]:
        print(f"    {r['telefono']}  usato da {r['n']} clienti (es. {r['clienti'][:60]}...)")

    if args.excel:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Telefoni da bonificare"
        ws.append(["Problema", "Dettaglio", "Cliente", "Telefono in anagrafica", "Email", "Telefono corretto (da compilare)"])
        for cella in ws[1]:
            cella.font = Font(bold=True)
        per_num_condiviso = {r["telefono"]: r["n"] for r in condivisi}
        for c, stato, numero, motivo in valutati:
            problema = None
            if stato in ("sospetto", "non_valido"):
                problema, dettaglio = stato.replace("_", " "), motivo
            elif stato == "mancante":
                problema, dettaglio = "mancante", ""
            elif (c["telefono"] or "") in per_num_condiviso:
                problema, dettaglio = "condiviso", f"stesso numero su {per_num_condiviso[c['telefono']]} clienti"
            if problema:
                ws.append([problema, dettaglio, c["nome"], c["telefono"] or "", c["email"] or "", ""])
        for i, larghezza in enumerate([14, 30, 34, 20, 28, 24], 1):
            ws.column_dimensions[get_column_letter(i)].width = larghezza
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:F{ws.max_row}"
        out = Path(args.excel) if isinstance(args.excel, str) else BASE / "liste" / f"telefoni_da_bonificare_{date.today().isoformat()}.xlsx"
        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out)
        print(f"\nLista di bonifica salvata in: {out} ({ws.max_row - 1} clienti)")


def cmd_archivia(args) -> None:
    conn = database.apri_db(Path(args.db))
    if args.ripristina:
        esito = database.archivia_pre(conn, ripristina=True)
        print(f"Ripristinati {esito['ripristinati']} veicoli archiviati.")
        return
    esito = database.archivia_pre(conn, cutoff=args.prima_di)
    print(f"Archiviati {esito['archiviati']} veicoli (immatricolati prima del {args.prima_di} "
          f"o senza data, e senza revisioni registrate).")
    print("Sono fuori da scadenzario, liste e 'da recuperare'; restano visibili nello storico cliente.")


def cmd_reset(args) -> None:
    percorso = Path(args.db)
    if not args.conferma:
        sys.exit("ATTENZIONE: 'reset' cancella TUTTI i dati (clienti, veicoli, storico).\n"
                 "Se sei sicuro, rilancia con --conferma. Aggiungi --demo per ricaricare dati fittizi.\n"
                 f"(Prima viene salvato un backup automatico in {percorso.parent}/backup)")
    database.esegui_backup(percorso)  # rete di sicurezza prima di cancellare
    conn = database.apri_db(percorso)
    database.svuota_dati(conn)
    print("Database svuotato (backup salvato in dati/backup).")
    if args.demo:
        from . import demo
        esito = demo.carica(conn)
        print(f"Dati demo caricati: {esito['clienti']} clienti, {esito['veicoli']} veicoli fittizi.")


def cmd_unisci_omonimi(args) -> None:
    conn = database.apri_db(Path(args.db))
    esito = database.unisci_omonimi(conn)
    print(f"Fusi {esito['clienti_fusi']} clienti duplicati in {esito['gruppi']} gruppi di omonimi.")
    print(f"Clienti totali ora: {conn.execute('SELECT COUNT(*) AS n FROM clienti').fetchone()['n']}")


def cmd_flotte(args) -> None:
    conn = database.apri_db(Path(args.db))
    if args.cliente:
        filtro = f"%{args.cliente}%"
        righe = conn.execute("SELECT * FROM clienti WHERE nome LIKE ?", (filtro,)).fetchall()
        if len(righe) != 1:
            for r in righe[:10]:
                print(f"  {r['nome']}  tel: {r['telefono'] or '-'}")
            sys.exit(f"{'Nessun cliente' if not righe else 'Più clienti'} per '{args.cliente}': restringi la ricerca.")
        database.imposta_flotta(conn, righe[0]["id"], not args.rimuovi)
        azione = "rimosso dalle flotte (torna nelle liste)" if args.rimuovi else "marcato come FLOTTA (escluso da SMS e chiamate)"
        print(f"{righe[0]['nome']}: {azione}.")
        return
    candidati = database.candidati_flotta(conn, soglia=args.soglia)
    if not candidati:
        print(f"Nessun cliente con almeno {args.soglia} veicoli attivi.")
        return
    print(f"Clienti con almeno {args.soglia} veicoli attivi (candidati flotta):\n")
    for r in candidati:
        stato = "FLOTTA" if r["flotta"] else "attivo nelle liste"
        print(f"  {r['nome']:<40} {r['veicoli']:>3} veicoli  [{stato}]")
    if args.applica:
        import re as re_mod
        azienda = re_mod.compile(r"\b(SRL|S\.R\.L|SPA|S\.P\.A|SNC|SAS|S\.A\.S|SCARL|COMUNE|RENTAL|FLOTTE|LEASING|NOLEGGIO|SOCIETA|AUTONOLEGG)\b")
        nuovi, dubbi = [], []
        for r in candidati:
            if r["flotta"]:
                continue
            if azienda.search(r["nome"]) or r["veicoli"] >= max(args.soglia * 2, 8):
                nuovi.append(r)
            else:
                dubbi.append(r)
        for r in nuovi:
            database.imposta_flotta(conn, r["id"], True)
        print(f"\nMarcati come flotta: {len(nuovi)} clienti (società evidenti o 8+ veicoli). "
              f"Esclusi da SMS e liste chiamate.")
        if dubbi:
            print("Da valutare a mano (privati con pochi veicoli, non toccati):")
            for r in dubbi:
                print(f"  flotte \"{r['nome']}\"   # {r['veicoli']} veicoli")
    else:
        print(f"\nUsa --applica per marcarli tutti come flotta, oppure "
              f"'flotte \"NOME\"' per gestirli uno a uno ('--rimuovi' per riattivare).")


def cmd_puntivendita(args) -> None:
    conn = database.apri_db(Path(args.db))
    righe = database.elenca_punti_vendita(conn)
    if not righe:
        print("Nessun punto vendita registrato: importa prima un file.")
        return
    print(f"{'COD':<5} {'DESCRIZIONE':<30} {'VEICOLI':>8}  STATO")
    print("-" * 60)
    for r in righe:
        stato = "ESCLUSO" if r["escluso"] else "attivo"
        print(f"{r['codice']:<5} {(r['descrizione'] or '-'):<30} {r['veicoli']:>8}  {stato}")


def _cmd_esclusione(args, escluso: bool) -> None:
    conn = database.apri_db(Path(args.db))
    pv = database.trova_punto_vendita(conn, args.punto_vendita)
    if not pv:
        sys.exit(f"Punto vendita non trovato (o ambiguo): '{args.punto_vendita}'. "
                 f"Usa 'puntivendita' per vedere codici e descrizioni.")
    database.imposta_esclusione(conn, pv["codice"], escluso)
    azione = "escluso da" if escluso else "reincluso in"
    print(f"Punto vendita {pv['codice']} {pv['descrizione'] or ''} {azione} scadenzario e liste.")


def cmd_stato(args) -> None:
    conn = database.apri_db(Path(args.db))
    veicoli = conn.execute("SELECT COUNT(*) AS n FROM veicoli").fetchone()["n"]
    clienti = conn.execute("SELECT COUNT(*) AS n FROM clienti").fetchone()["n"]
    print(f"Database: {args.db}")
    print(f"Veicoli: {veicoli} — Clienti: {clienti}")
    print("Ultimi import:")
    for r in conn.execute("SELECT * FROM import_log ORDER BY id DESC LIMIT 5"):
        print(f"  {r['data_import']}  {r['file']}  ({r['righe_lette']} righe, {r['veicoli_nuovi']} nuovi)")


def main(argv=None) -> None:
    parser = argparse.ArgumentParser(prog="revisioni", description="Gestione automatica scadenzario revisioni.")
    parser.add_argument("--db", default=str(DB_PREDEFINITO), help="Percorso del database (predefinito: dati/revisioni.db)")
    sub = parser.add_subparsers(dest="comando", required=True)

    p = sub.add_parser("importa", help="Importa un riepilogo mensile C.C.F. di WinDrakkar (xlsx)")
    p.add_argument("file")
    p.set_defaults(func=cmd_importa)

    p = sub.add_parser("importa-storico", help="Importa file o cartelle dello storico (formato Cristina, .xlsx/.xls)")
    p.add_argument("percorso", help="File singolo o cartella (ricorsiva)")
    p.set_defaults(func=cmd_importa_storico)

    p = sub.add_parser("importa-immatricolazioni", help="Importa il registro immatricolazioni del collega (date autorevoli)")
    p.add_argument("percorso", help="Cartella (ricorsiva) o file singolo")
    p.set_defaults(func=cmd_importa_immatricolazioni)

    p = sub.add_parser("importa-dekra", help="Importa l'export PDF del portale Dekra (revisioni fatte in officina)")
    p.add_argument("percorso", help="File PDF o cartella")
    p.set_defaults(func=cmd_importa_dekra)

    p = sub.add_parser("importa-tcar", help="Importa un export lead di Tcar (EstrazioneLeads*.xlsx)")
    p.add_argument("file")
    p.set_defaults(func=cmd_importa_tcar)

    p = sub.add_parser("lead", help="Elenca i lead Tcar agganciati ai veicoli nostri")
    p.add_argument("--tutti", action="store_true", help="Mostra anche i lead non agganciati")
    p.set_defaults(func=cmd_lead)

    p = sub.add_parser("scadenze", help="Mostra i veicoli da contattare")
    p.add_argument("--giorni", type=int, default=60, help="Orizzonte allarme in giorni (predefinito 60)")
    p.add_argument("--tutte", action="store_true", help="Mostra tutto il parco, anche scadenze future")
    p.add_argument("--limite", type=int, default=None, help="Numero massimo di righe a video")
    p.add_argument("--includi-esclusi", action="store_true", help="Considera anche i punti vendita esclusi")
    p.add_argument("--mese", help="Tutte le scadenze di un mese, anche futuro (AAAA-MM, es. 2026-09)")
    p.set_defaults(func=cmd_scadenze)

    p = sub.add_parser("excel", help="Genera la lista chiamate in Excel")
    p.add_argument("--giorni", type=int, default=60)
    p.add_argument("--tutte", action="store_true")
    p.add_argument("--out", help="File di destinazione")
    p.add_argument("--includi-esclusi", action="store_true", help="Considera anche i punti vendita esclusi")
    p.set_defaults(func=cmd_excel)

    p = sub.add_parser("sms", help="Estrae la lista SMS per smscafè (scadenze tra 2 mesi)")
    p.add_argument("--arretrato", action="store_true", help="Campagna di recupero su tutto l'arretrato")
    p.add_argument("--segna", action="store_true", help="Registra l'esito 'sms_inviato' sui veicoli estratti")
    p.add_argument("--out", help="File CSV di destinazione")
    p.set_defaults(func=cmd_sms)

    p = sub.add_parser("archivia", help="Archivia i veicoli non gestibili (pre-cutoff o senza data, senza revisioni)")
    p.add_argument("--prima-di", default="2022-01-01", help="Data di taglio ISO (default 2022-01-01)")
    p.add_argument("--ripristina", action="store_true", help="Riporta in gestione tutti gli archiviati")
    p.set_defaults(func=cmd_archivia)

    p = sub.add_parser("reset", help="Svuota il database (--demo per caricare dati fittizi di prova)")
    p.add_argument("--conferma", action="store_true", help="Conferma la cancellazione di tutti i dati")
    p.add_argument("--demo", action="store_true", help="Dopo lo svuotamento carica dati fittizi di demo")
    p.set_defaults(func=cmd_reset)

    p = sub.add_parser("unisci-omonimi", help="Fonde i clienti duplicati con nome identico (telefoni diversi)")
    p.set_defaults(func=cmd_unisci_omonimi)

    p = sub.add_parser("flotte", help="Rileva/gestisce le flotte (clienti con molti veicoli, esclusi dalle liste)")
    p.add_argument("cliente", nargs="?", help="Nome cliente da marcare/smarcare manualmente")
    p.add_argument("--soglia", type=int, default=4, help="Veicoli attivi minimi per candidatura (default 4)")
    p.add_argument("--applica", action="store_true", help="Marca come flotta tutti i candidati")
    p.add_argument("--rimuovi", action="store_true", help="Con nome cliente: toglie il flag flotta")
    p.set_defaults(func=cmd_flotte)

    p = sub.add_parser("telefoni", help="Report qualità telefoni (fittizi, non validi, condivisi)")
    p.add_argument("--excel", nargs="?", const=True, help="Genera la lista Excel di bonifica (percorso opzionale)")
    p.set_defaults(func=cmd_telefoni)

    p = sub.add_parser("puntivendita", help="Elenca punti vendita/venditori e il loro stato")
    p.set_defaults(func=cmd_puntivendita)

    p = sub.add_parser("escludi", help="Escludi un punto vendita dalle liste (codice o nome, es. 48 o livio)")
    p.add_argument("punto_vendita")
    p.set_defaults(func=lambda a: _cmd_esclusione(a, True))

    p = sub.add_parser("includi", help="Reincludi un punto vendita nelle liste")
    p.add_argument("punto_vendita")
    p.set_defaults(func=lambda a: _cmd_esclusione(a, False))

    p = sub.add_parser("esito", help="Registra l'esito di un contatto (per telaio o targa)")
    p.add_argument("veicolo", help="Telaio o targa")
    p.add_argument("esito", choices=["contattato", "appuntamento", "non_interessato",
                                     "irraggiungibile", "revisione_fatta", "non_possiede_piu"])
    p.add_argument("--data", help="Data revisione effettuata (AAAA-MM-GG), solo per revisione_fatta")
    p.add_argument("--note")
    p.set_defaults(func=cmd_esito)

    p = sub.add_parser("cliente", help="Storico completo di un cliente (per nome o telefono)")
    p.add_argument("ricerca")
    p.set_defaults(func=cmd_cliente)

    p = sub.add_parser("nuovo-veicolo", help="Registra il nuovo veicolo di un cliente (comunicato a voce)")
    p.add_argument("cliente", help="Nome (anche parziale) o telefono del cliente")
    p.add_argument("--targa")
    p.add_argument("--telaio")
    p.add_argument("--marca")
    p.add_argument("--modello")
    p.add_argument("--imm", help="Data immatricolazione AAAA-MM-GG")
    p.add_argument("--ultima-revisione", help="Data ultima revisione AAAA-MM-GG (per usati)")
    p.add_argument("--telefono", help="Necessario solo se il cliente è nuovo")
    p.set_defaults(func=cmd_nuovo_veicolo)

    p = sub.add_parser("da-recuperare", help="Clienti storici senza veicolo in scadenzario")
    p.set_defaults(func=cmd_da_recuperare)

    p = sub.add_parser("storico", help="Genera lo storico clienti navigabile (HTML)")
    p.add_argument("--out", help="File di destinazione")
    p.set_defaults(func=cmd_storico)

    p = sub.add_parser("stato", help="Riepilogo del database")
    p.set_defaults(func=cmd_stato)

    p = sub.add_parser("web", help="Avvia la dashboard web locale")
    p.add_argument("--porta", type=int, default=8765)
    p.add_argument("--no-browser", action="store_true", help="Non aprire il browser automaticamente")
    p.add_argument("--rete", action="store_true", help="Raggiungibile anche dagli altri PC della rete")
    p.set_defaults(func=lambda a: __import__("revisioni.web", fromlist=["avvia"]).avvia(
        Path(a.db), a.porta, not a.no_browser, a.rete))

    p = sub.add_parser("backup", help="Esegue subito il backup del database")
    p.set_defaults(func=lambda a: [print(f"Backup → {e}") for e in database.esegui_backup(Path(a.db))])

    args = parser.parse_args(argv)
    args.func(args)


if __name__ == "__main__":
    main()
