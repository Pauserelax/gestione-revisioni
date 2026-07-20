"""Esportazione delle liste di chiamata in Excel."""

from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .scadenze import ScadenzaVeicolo

COLONNE = [
    ("Stato", 14), ("Scadenza revisione", 18), ("Giorni", 8),
    ("Cliente", 32), ("Telefono", 18), ("Targa", 10), ("Marca", 8), ("Modello", 10),
    ("Telaio", 20), ("Immatricolazione", 16), ("Fonte data", 12),
    ("P.V.", 6), ("Lead Tcar", 30), ("Avviso", 34), ("Fonte dati", 30),
    ("Ultimo esito", 14), ("Note operatore", 40),
]

RIEMPIMENTI = {
    "SCADUTA": PatternFill("solid", fgColor="F4CCCC"),
    "IN_SCADENZA": PatternFill("solid", fgColor="FFF2CC"),
    "PROSSIMA": PatternFill("solid", fgColor="D9EAD3"),
}


def esporta_lista_chiamate(scadenze: list[ScadenzaVeicolo], percorso: Path, oggi: date | None = None) -> Path:
    oggi = oggi or date.today()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista chiamate revisioni"

    ws.append([f"Lista chiamate revisioni — generata il {oggi.strftime('%d/%m/%Y')}"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([c[0] for c in COLONNE])
    for cella in ws[2]:
        cella.font = Font(bold=True)

    for s in scadenze:
        ws.append([
            s.stato,
            s.scadenza.strftime("%d/%m/%Y") if s.scadenza else "",
            s.giorni_rimanenti if s.giorni_rimanenti is not None else "",
            s.cliente,
            s.telefono,
            s.targa,
            s.marca,
            s.modello,
            s.telaio,
            s.data_immatricolazione.strftime("%d/%m/%Y") if s.data_immatricolazione else "",
            "presunta" if s.fonte_data == "mese_report" else s.fonte_data,
            s.punto_vendita,
            s.lead_tcar,
            s.avviso,
            s.fonte_file,
            s.ultimo_esito,
            "",
        ])
        riempimento = RIEMPIMENTI.get(s.stato)
        if riempimento:
            ws.cell(row=ws.max_row, column=1).fill = riempimento

    for i, (_, larghezza) in enumerate(COLONNE, 1):
        ws.column_dimensions[get_column_letter(i)].width = larghezza
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLONNE))}{ws.max_row}"

    percorso.parent.mkdir(parents=True, exist_ok=True)
    wb.save(percorso)
    return percorso
