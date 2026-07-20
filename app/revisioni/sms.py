"""Generazione delle liste SMS per smscafè.

Flusso: primo contatto via SMS ai veicoli in scadenza tra 2 mesi; campagna di
recupero via SMS generico su tutto l'arretrato. Il testo è personalizzabile nei
file dati/sms_testo.txt e dati/sms_testo_arretrato.txt (creati al primo uso).
"""

from __future__ import annotations

import csv
import re
import sqlite3
from datetime import date
from pathlib import Path

from . import db as database
from .scadenze import ScadenzaVeicolo, calcola_scadenze, da_contattare

MESI_IT = ["", "gennaio", "febbraio", "marzo", "aprile", "maggio", "giugno",
           "luglio", "agosto", "settembre", "ottobre", "novembre", "dicembre"]

# Struttura richiesta: nome del cliente, identificazione del veicolo,
# mittente con contatto. Massimo 160 caratteri (1 SMS).
MAX_SMS = 160
TESTO_PREDEFINITO = (
    "Gentile {NOME}, la revisione della sua {MODELLO} targa {TARGA} scade "
    "entro {MESE}. Per prenotare: Centro Car Cazzaro {TEL_OFFICINA}"
)
TESTO_ARRETRATO_PREDEFINITO = (
    "Gentile {NOME}, dai nostri archivi la revisione della sua auto potrebbe "
    "essere scaduta. Verifichi con noi: Centro Car Cazzaro {TEL_OFFICINA}"
)


def _testo_template(percorso: Path, predefinito: str) -> str:
    if not percorso.exists():
        percorso.parent.mkdir(parents=True, exist_ok=True)
        percorso.write_text(predefinito, encoding="utf-8")
    return percorso.read_text(encoding="utf-8").strip()


from .telefoni import cellulare_valido as _cellulare


def _sostituisci(testo: str, nome: str, modello: str, s: ScadenzaVeicolo) -> str:
    mese = f"{MESI_IT[s.scadenza.month]} {s.scadenza.year}" if s.scadenza else ""
    msg = (testo.replace("{NOME}", nome)
               .replace("{MARCA}", "")
               .replace("{MODELLO}", modello)
               .replace("{TARGA}", s.targa)
               .replace("{MESE}", mese))
    return re.sub(r"\s{2,}", " ", msg).replace(" ,", ",").strip()


def _modello_leggibile(modello: str) -> str:
    from .modelli import normalizza_modello
    nome = normalizza_modello(modello)
    # codice interno non mappato (corto, con cifre): meglio "auto" che una sigla
    if len(nome) <= 4 and any(ch.isdigit() for ch in nome):
        return "auto"
    return (nome or "auto").title()


def _componi(testo: str, s: ScadenzaVeicolo) -> str:
    """Compone il messaggio restando entro MAX_SMS caratteri.

    Se il testo pieno sfora, accorcia per gradi: prima toglie il modello
    (resta la targa a identificare il veicolo), poi riduce il nome alle
    prime due parole. Il nome del cliente e il mittente non si toccano."""
    nome_pieno = s.cliente.title()
    nome_corto = " ".join(nome_pieno.split()[:2])
    varianti = [
        (nome_pieno, _modello_leggibile(s.modello)),
        (nome_pieno, "auto"),
        (nome_corto, "auto"),
    ]
    msg = ""
    for nome, modello in varianti:
        msg = _sostituisci(testo, nome, modello, s)
        if len(msg) <= MAX_SMS:
            return msg
    return msg


def controlla_template(testo: str) -> None:
    """Blocca l'estrazione se nel testo restano segnaposto non compilati
    (tipicamente {TEL_OFFICINA}: il numero va scritto in dati/sms_testo.txt)."""
    residui = set(re.findall(r"\{[A-Z_]+\}", testo)) - {"{NOME}", "{MARCA}", "{MODELLO}", "{TARGA}", "{MESE}"}
    if residui:
        raise ValueError(
            f"Testo SMS incompleto: sostituisci {', '.join(sorted(residui))} nel file dei testi "
            f"(dati/sms_testo.txt) con il dato reale, es. il telefono dell'officina.")


def lista_sms(conn: sqlite3.Connection, arretrato: bool = False,
              base: Path | None = None, oggi: date | None = None) -> list[dict]:
    """Righe SMS pronte per smscafè. Fase SMS (+2 mesi) o campagna arretrato.

    Un solo SMS per numero di telefono; esclusi i veicoli senza cellulare
    valido e quelli con esito già registrato per la scadenza corrente."""
    base = base or Path(".")
    oggi = oggi or date.today()
    scadenze = da_contattare(calcola_scadenze(conn, oggi=oggi))
    if arretrato:
        selezione = [s for s in scadenze if s.fase == "ARRETRATO" and s.ultimo_esito == ""]
        testo = _testo_template(base / "dati" / "sms_testo_arretrato.txt", TESTO_ARRETRATO_PREDEFINITO)
    else:
        selezione = [s for s in scadenze if s.fase == "SMS" and s.ultimo_esito != "sms_inviato"]
        testo = _testo_template(base / "dati" / "sms_testo.txt", TESTO_PREDEFINITO)
    controlla_template(testo)

    righe, visti, senza_numero = [], set(), 0
    for s in selezione:
        numero = _cellulare(s.telefono)
        if not numero:
            senza_numero += 1
            continue
        if numero in visti:
            continue
        visti.add(numero)
        righe.append({
            "telefono": numero,
            "nome": s.cliente,
            "targa": s.targa,
            "marca": s.marca,
            "modello": s.modello,
            "scadenza": s.scadenza.strftime("%d/%m/%Y") if s.scadenza else "",
            "messaggio": _componi(testo, s),
            "veicolo_id": s.veicolo_id,
            "scadenza_iso": s.scadenza.isoformat() if s.scadenza else "",
        })
    righe.sort(key=lambda r: (r["scadenza"], r["nome"]))
    return righe


def _dividi_nome(nominativo: str) -> tuple[str, str]:
    """WinDrakkar scrive COGNOME NOME: prima parola = cognome, resto = nome."""
    parti = (nominativo or "").split()
    if len(parti) <= 1:
        return nominativo or "", ""
    return parti[0], " ".join(parti[1:])


def testo_campagna_smscafe(testo_template: str, righe: list[dict]) -> str:
    """Il testo unico da incollare in SMS Café: {NOME} diventa il campo <Nome>,
    il veicolo diventa "auto" (il campo unione copre solo i dati di rubrica) e
    {MESE} è il mese della campagna."""
    from collections import Counter
    mese_txt = ""
    mesi = Counter(r["scadenza"][3:] for r in righe if r.get("scadenza") and len(r["scadenza"]) == 10)
    if mesi:
        mm_aaaa = mesi.most_common(1)[0][0]          # "09/2026"
        mese_txt = f"{MESI_IT[int(mm_aaaa[:2])]} {mm_aaaa[3:]}"
    msg = (testo_template
           .replace("{MODELLO} targa {TARGA}", "auto")
           .replace("{MODELLO}", "auto")
           .replace("{TARGA}", "")
           .replace("{NOME}", "<Nome>")
           .replace("{MESE}", mese_txt))
    return re.sub(r"\s{2,}", " ", msg).replace(" ,", ",").strip()


def _nome_entro_budget(nominativo: str, budget: int) -> str:
    """Il valore per la colonna Nome, garantito entro il budget di caratteri:
    nomi di battesimo completi → solo il primo nome → "Cliente" (aziende o
    budget risicato). Così OGNI messaggio resta entro i 160 caratteri."""
    cognome, nome = _dividi_nome(nominativo)
    base = (nome or cognome or "").title()
    candidati = [base]
    if base:
        candidati.append(base.split()[0])
    candidati.append("Cliente")
    for c in candidati:
        if c and len(c) <= budget:
            return c
    return "Cliente"


def crea_excel_smscafe(righe: list[dict], testo_template: str | None = None) -> bytes:
    """File Excel nel formato atteso dall'import di SMS Café:
    colonna A = Nome (nome proprio: è quello che compila il campo <Nome>),
    B = Cognome, C = Cellulare (obbligatorio), prima riga con etichette.
    Il secondo foglio contiene il testo della campagna pronto da incollare."""
    import io

    import openpyxl
    from openpyxl.styles import Font
    # Budget caratteri per il nome: 160 meno la parte fissa del testo campagna.
    testo = testo_campagna_smscafe(testo_template, righe) if testo_template else ""
    budget_nome = MAX_SMS - (len(testo) - len("<Nome>")) if testo else 24

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contatti"
    ws.append(["Nome", "Cognome", "Cellulare", "Targa", "Scadenza revisione", "Messaggio singolo (riferimento)"])
    for cella in ws[1]:
        cella.font = Font(bold=True)
    generici = 0
    nome_piu_lungo = 0
    for r in righe:
        cognome, _ = _dividi_nome(r["nome"])
        nome_proprio = _nome_entro_budget(r["nome"], budget_nome)
        if nome_proprio == "Cliente":
            generici += 1
        nome_piu_lungo = max(nome_piu_lungo, len(nome_proprio))
        r["nome_proprio"] = nome_proprio
        ws.append([nome_proprio, cognome.title(), r["telefono"], r.get("targa", ""),
                   r.get("scadenza", ""), r.get("messaggio", "")])
    for lettera, larghezza in zip("ABCDEF", (22, 22, 14, 10, 16, 70)):
        ws.column_dimensions[lettera].width = larghezza

    if testo:
        lunghezza_max = len(testo) - len("<Nome>") + max(nome_piu_lungo, len("Cliente"))
        ws2 = wb.create_sheet("Testo campagna")
        ws2["A1"] = "TESTO DA INCOLLARE NEL MESSAGGIO DI SMS CAFÉ"
        ws2["A1"].font = Font(bold=True)
        ws2["A3"] = testo
        ws2["A5"] = "<Nome> viene sostituito da SMS Café col nome del contatto."
        ws2["A6"] = (f"Lunghezza massima garantita: {lunghezza_max} caratteri "
                     + ("(OK, entro 160: i nomi troppo lunghi sono già stati accorciati nella colonna Nome)"
                        if lunghezza_max <= MAX_SMS else "(ATTENZIONE: accorcia il testo della campagna)"))
        if generici:
            ws2["A7"] = f'{generici} contatti senza nome utilizzabile (aziende/nomi lunghissimi): riceveranno "Gentile Cliente".'
        ws2["A8"] = "Import: 1ª riga etichette = SÌ; Nome=Colonna A, Cognome=Colonna B, Cellulare=Colonna C."
        ws2["A9"] = "Consiglio: importa in una categoria dedicata (es. REVISIONI SETTEMBRE 2026)."
        ws2.column_dimensions["A"].width = 120
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def esporta_smscafe(righe: list[dict], percorso: Path, testo_template: str | None = None) -> Path:
    percorso.parent.mkdir(parents=True, exist_ok=True)
    percorso.write_bytes(crea_excel_smscafe(righe, testo_template))
    return percorso


def esporta_csv(righe: list[dict], percorso: Path) -> Path:
    percorso.parent.mkdir(parents=True, exist_ok=True)
    with open(percorso, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["Telefono", "Nome", "Targa", "Scadenza", "Messaggio", "Caratteri"])
        for r in righe:
            w.writerow([r["telefono"], r["nome"], r["targa"], r["scadenza"], r["messaggio"], len(r["messaggio"])])
    return percorso


def segna_inviati(conn: sqlite3.Connection, righe: list[dict]) -> int:
    """Registra l'esito 'sms_inviato' per i veicoli della lista."""
    for r in righe:
        database.registra_contatto(conn, r["veicolo_id"], r["scadenza_iso"], "sms_inviato",
                                   "estrazione lista smscafè")
    return len(righe)
